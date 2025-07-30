#!/usr/bin/env python3
"""
Complete demonstration of the tidyxl package

This comprehensive example showcases all tidyxl functions with practical use cases:
- xlsx_cells(): Extract cell data in tidy format
- xlsx_sheet_names(): List worksheet names
- xlsx_names(): Extract named ranges
- xlsx_validation(): Extract data validation rules
- xlsx_formats(): Extract formatting information

The demo creates a complex Excel file with various features, then shows how to
analyze it using tidyxl functions.
"""

import pandas as pd
from tidyxl import xlsx_cells, xlsx_sheet_names, xlsx_names, xlsx_validation, xlsx_formats
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from datetime import datetime
import os

print("=" * 80)
print("TIDYXL COMPLETE DEMONSTRATION")
print("Python package for tidy Excel data extraction")
print("=" * 80)

# =============================================================================
# PART 1: CREATE COMPLEX SAMPLE EXCEL FILE
# =============================================================================

print("\n1. Creating comprehensive sample Excel file...")

wb = Workbook()

# --- SHEET 1: SALES DATA WITH VARIOUS DATA TYPES ---
ws1 = wb.active
ws1.title = "Sales_Data"

# Add headers with formatting
headers = ["Product", "Region", "Sales_Amount", "Target", "Met_Target", "Sale_Date", "Notes"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = cell.font.copy(bold=True)

# Add sample data with various types
sales_data = [
    ["Widget A", "North", 15000.50, 12000, True, datetime(2023, 1, 15), "Strong performance"],
    ["Widget B", "South", 8500.25, 10000, False, datetime(2023, 1, 20), "Needs improvement"],
    ["Widget C", "East", 22000.00, 20000, True, datetime(2023, 1, 25), "Exceeded target"],
    ["Widget D", "West", 12750.75, 15000, False, datetime(2023, 2, 1), ""],
    ["Widget E", "North", 18500.00, 16000, True, datetime(2023, 2, 5), "Great quarter"]
]

for row_idx, row_data in enumerate(sales_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        
        # Add a comment to one cell
        if row_idx == 2 and col_idx == 1:
            cell.comment = Comment("This is our bestselling product", "Sales Manager")

# Add formulas
ws1["H1"] = "Total_Sales"
ws1["H2"] = "=SUM(C2:C6)"
ws1["I1"] = "Avg_Sales"
ws1["I2"] = "=AVERAGE(C2:C6)"

# --- SHEET 2: VALIDATION EXAMPLES ---
ws2 = wb.create_sheet("Data_Entry")

# Headers
validation_headers = ["Score", "Date", "Category", "Comments"]
for col, header in enumerate(validation_headers, 1):
    ws2.cell(row=1, column=col, value=header)

# Add validation rules
# 1. Number validation (scores 0-100)
score_validation = DataValidation(
    type="whole",
    operator="between", 
    formula1=0,
    formula2=100,
    showErrorMessage=True,
    errorTitle="Invalid Score",
    error="Score must be between 0 and 100"
)
ws2.add_data_validation(score_validation)
score_validation.add("A2:A20")

# 2. Date validation
date_validation = DataValidation(
    type="date",
    operator="greaterThan",
    formula1="2020-01-01",
    showErrorMessage=True,
    errorTitle="Invalid Date", 
    error="Date must be after 2020-01-01"
)
ws2.add_data_validation(date_validation)
date_validation.add("B2:B20")

# 3. List validation
category_validation = DataValidation(
    type="list",
    formula1='"Excellent,Good,Average,Poor"',
    showErrorMessage=True,
    errorTitle="Invalid Category",
    error="Please select from the dropdown list"
)
ws2.add_data_validation(category_validation)
category_validation.add("C2:C20")

# --- SHEET 3: SUMMARY WITH NAMED RANGES ---
ws3 = wb.create_sheet("Summary")

# Summary data
ws3["A1"] = "Sales Summary Report"
ws3["A3"] = "Total Products:"
ws3["B3"] = "=COUNTA(Sales_Data.A2:A6)"
ws3["A4"] = "Total Revenue:"
ws3["B4"] = "=SUM(Sales_Range)"
ws3["A5"] = "Average Sale:"
ws3["B5"] = "=AVERAGE(Sales_Range)"

# Add named ranges
wb.defined_names["Sales_Range"] = DefinedName("Sales_Range", attr_text="Sales_Data.C2:C6")
wb.defined_names["Product_List"] = DefinedName("Product_List", attr_text="Sales_Data.A2:A6")
wb.defined_names["Summary_Title"] = DefinedName("Summary_Title", attr_text="Summary.A1", localSheetId=2)

# Save the workbook
filename = "tidyxl_demo.xlsx"
wb.save(filename)
wb.close()

print(f"✓ Created '{filename}' with 3 worksheets:")
print("  - Sales_Data: Various data types, formulas, comments")
print("  - Data_Entry: Data validation examples")
print("  - Summary: Named ranges and summary formulas")

# =============================================================================
# PART 2: DEMONSTRATE ALL TIDYXL FUNCTIONS
# =============================================================================

print("\n" + "=" * 80)
print("2. DEMONSTRATING TIDYXL FUNCTIONS")
print("=" * 80)

# --- FUNCTION 1: xlsx_sheet_names() ---
print("\n📋 xlsx_sheet_names() - List all worksheet names")
print("-" * 50)

sheets = xlsx_sheet_names(filename)
print(f"Found {len(sheets)} worksheets:")
for i, sheet in enumerate(sheets, 1):
    print(f"  {i}. {sheet}")

# --- FUNCTION 2: xlsx_names() ---
print("\n🏷️  xlsx_names() - Extract named ranges")
print("-" * 40)

names_df = xlsx_names(filename)
print(f"Found {len(names_df)} named ranges:")
if len(names_df) > 0:
    print(names_df[['name', 'formula', 'sheet', 'is_range']].to_string(index=False))
    
    print(f"\nBreakdown:")
    global_count = len(names_df[names_df['sheet'].isna()])
    sheet_count = len(names_df[names_df['sheet'].notna()])
    print(f"  Global ranges: {global_count}")
    print(f"  Sheet-specific ranges: {sheet_count}")

# --- FUNCTION 3: xlsx_validation() ---
print("\n✅ xlsx_validation() - Extract data validation rules")
print("-" * 52)

validation_df = xlsx_validation(filename)
print(f"Found {len(validation_df)} validation rules:")
if len(validation_df) > 0:
    # Show key validation info
    key_cols = ['sheet', 'ref', 'type', 'operator', 'formula1', 'error_title']
    display_cols = [col for col in key_cols if col in validation_df.columns]
    print(validation_df[display_cols].to_string(index=False))
    
    print(f"\nValidation types:")
    for vtype, count in validation_df['type'].value_counts().items():
        print(f"  {vtype}: {count} rules")

# --- FUNCTION 4: xlsx_formats() ---
print("\n🎨 xlsx_formats() - Extract formatting information")
print("-" * 48)

formats = xlsx_formats(filename)
print("Formatting categories found:")
for category, items in formats.items():
    print(f"  {category}: {len(items)} entries")

if formats['fonts']:
    print(f"\nExample font: {formats['fonts'][0]}")

# --- FUNCTION 5: xlsx_cells() - THE MAIN FUNCTION ---
print("\n📊 xlsx_cells() - Extract all cell data in tidy format")
print("-" * 54)

# Read all cells
all_cells = xlsx_cells(filename)
print(f"Total cells extracted: {len(all_cells)}")
print(f"Sheets processed: {all_cells['sheet'].unique().tolist()}")

# Show data type distribution
print(f"\nData type distribution:")
type_counts = all_cells['data_type'].value_counts()
for dtype, count in type_counts.items():
    print(f"  {dtype}: {count} cells")

# Show first few content cells
content_cells = all_cells[~all_cells['is_blank']].head(10)
print(f"\nFirst 10 cells with content:")
display_cols = ['sheet', 'address', 'data_type', 'content']
print(content_cells[display_cols].to_string(index=False))

# =============================================================================
# PART 3: PRACTICAL ANALYSIS EXAMPLES
# =============================================================================

print("\n" + "=" * 80)
print("3. PRACTICAL ANALYSIS EXAMPLES")
print("=" * 80)

# --- EXAMPLE 1: FIND ALL FORMULAS ---
print("\n🧮 Example 1: Finding all formulas")
print("-" * 35)

formulas = all_cells[all_cells['formula'].notna()]
print(f"Found {len(formulas)} formula cells:")
for _, cell in formulas.iterrows():
    print(f"  {cell['address']}: {cell['formula']}")

# --- EXAMPLE 2: ANALYZE NUMERIC DATA ---
print("\n📈 Example 2: Analyzing numeric data")
print("-" * 37)

numeric_cells = all_cells[all_cells['data_type'] == 'numeric']
if len(numeric_cells) > 0:
    values = numeric_cells['numeric'].dropna()
    print(f"Numeric cells: {len(numeric_cells)}")
    print(f"  Min value: {values.min():,.2f}")
    print(f"  Max value: {values.max():,.2f}")
    print(f"  Average: {values.mean():,.2f}")

# --- EXAMPLE 3: FIND COMMENTS ---
print("\n💬 Example 3: Finding cell comments")
print("-" * 34)

comments = all_cells[all_cells['comment'].notna()]
print(f"Found {len(comments)} cells with comments:")
for _, cell in comments.iterrows():
    print(f"  {cell['address']}: {cell['comment'][:50]}{'...' if len(str(cell['comment'])) > 50 else ''}")

# --- EXAMPLE 4: SHEET-SPECIFIC ANALYSIS ---
print("\n📋 Example 4: Sheet-specific analysis")
print("-" * 36)

sales_data = all_cells[all_cells['sheet'] == 'Sales_Data']
sales_content = sales_data[~sales_data['is_blank']]
print(f"Sales_Data sheet: {len(sales_content)} cells with content")

# Find headers (row 1)
headers = sales_content[sales_content['row'] == 1]
print(f"Headers found: {', '.join(headers['character'].dropna().tolist())}")

# --- EXAMPLE 5: CONVERT BACK TO TABULAR FORMAT ---
print("\n🔄 Example 5: Converting tidy data back to table")
print("-" * 47)

def tidy_to_table(cells_df, sheet_name, max_rows=5):
    """Convert tidy format back to tabular format for display"""
    sheet_cells = cells_df[(cells_df['sheet'] == sheet_name) & (~cells_df['is_blank'])]
    if len(sheet_cells) == 0:
        return pd.DataFrame()
    
    # Limit to first few rows for display
    limited_cells = sheet_cells[sheet_cells['row'] <= max_rows]
    
    return limited_cells.pivot_table(
        index='row',
        columns='col', 
        values='content',
        aggfunc='first'
    ).fillna('')

table = tidy_to_table(all_cells, 'Sales_Data')
if not table.empty:
    print("Sales_Data reconstructed (first 5 rows):")
    print(table.to_string())

# =============================================================================
# PART 4: ADVANCED USE CASES
# =============================================================================

print("\n" + "=" * 80)
print("4. ADVANCED USE CASES")
print("=" * 80)

# --- USE CASE 1: DATA QUALITY ASSESSMENT ---
print("\n🔍 Use Case 1: Data quality assessment")
print("-" * 38)

# Find potential data quality issues
blank_in_data = all_cells[
    (all_cells['sheet'] == 'Sales_Data') & 
    (all_cells['row'] > 1) & 
    (all_cells['col'] <= 7) & 
    (all_cells['is_blank'])
]
print(f"Blank cells in data range: {len(blank_in_data)}")

# Check for mixed data types in columns
sales_cells = all_cells[all_cells['sheet'] == 'Sales_Data']
for col in range(1, 8):  # Check first 7 columns
    col_data = sales_cells[(sales_cells['col'] == col) & (sales_cells['row'] > 1)]
    types = col_data['data_type'].unique()
    if len(types) > 1:
        print(f"Column {col}: Mixed data types {types}")

# --- USE CASE 2: CROSS-SHEET REFERENCES ---
print("\n🔗 Use Case 2: Finding cross-sheet references")
print("-" * 43)

cross_refs = formulas[formulas['formula'].str.contains('Sales_Data', na=False)]
print(f"Formulas referencing other sheets: {len(cross_refs)}")
for _, cell in cross_refs.iterrows():
    print(f"  {cell['sheet']}.{cell['address']}: {cell['formula']}")

# --- USE CASE 3: VALIDATION COVERAGE ---
print("\n📊 Use Case 3: Validation rule coverage")
print("-" * 37)

if len(validation_df) > 0:
    total_validated_cells = 0
    for _, rule in validation_df.iterrows():
        ref = rule['ref']
        if ':' in ref:  # Range like A2:A20
            # Simple approximation for demo
            parts = ref.split(':')
            start_row = int(''.join(filter(str.isdigit, parts[0])))
            end_row = int(''.join(filter(str.isdigit, parts[1])))
            total_validated_cells += (end_row - start_row + 1)
    
    print(f"Total cells with validation rules: ~{total_validated_cells}")

print("\n" + "=" * 80)
print("DEMONSTRATION COMPLETE!")
print("=" * 80)

print(f"\nKey Benefits of tidyxl:")
print("• 📊 Tidy format: Each cell is one row with complete metadata")
print("• 🔍 Comprehensive: Extract values, formulas, formatting, comments")
print("• 🏷️  Named ranges: Access Excel defined names and ranges") 
print("• ✅ Validation: Discover data entry rules and constraints")
print("• 📋 Multi-sheet: Process all worksheets or specific ones")
print("• 🔄 Flexible: Perfect for complex, non-tabular Excel analysis")
print("• 🎯 R Compatible: Identical to R tidyxl package")

print(f"\nNext steps:")
print("• Explore your own Excel files using these functions")
print("• Combine with pandas for advanced data analysis")
print("• Use for data quality assessment and cleaning")
print("• Perfect for messy, real-world Excel spreadsheets")

# Clean up
if os.path.exists(filename):
    os.remove(filename)
    print(f"\n🗑️  Cleaned up demo file: {filename}")