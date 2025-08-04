#!/usr/bin/env python3
"""
Demonstration of the new tidyxl functions: xlsx_sheet_names, xlsx_names, xlsx_validation
"""

import pandas as pd
from tidyxl import xlsx_sheet_names, xlsx_names, xlsx_validation, xlsx_cells
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

print("=" * 80)
print("NEW TIDYXL FUNCTIONS DEMONSTRATION")
print("=" * 80)

# Create a comprehensive test Excel file with named ranges and validation
print("Creating test Excel file with named ranges and data validation...")

wb = Workbook()

# Sheet 1: Main data with named ranges
ws1 = wb.active
ws1.title = "Sales_Data"

# Add some sample data
data = [
    ["Product", "Region", "Sales", "Target"],
    ["Widget A", "North", 1000, 1200],
    ["Widget B", "South", 1500, 1400],
    ["Widget C", "East", 800, 900],
    ["Widget D", "West", 2000, 1800]
]

for row_idx, row_data in enumerate(data, 1):
    for col_idx, value in enumerate(row_data, 1):
        ws1.cell(row=row_idx, column=col_idx, value=value)

# Add named ranges
from openpyxl.workbook.defined_name import DefinedName
wb.defined_names["SalesData"] = DefinedName("SalesData", attr_text="Sales_Data.A1:D5")
wb.defined_names["ProductList"] = DefinedName("ProductList", attr_text="Sales_Data.A2:A5") 
wb.defined_names["SalesColumn"] = DefinedName("SalesColumn", attr_text="Sales_Data.C2:C5")

# Sheet 2: Validation examples
ws2 = wb.create_sheet("Validation_Examples")

# Add headers
validation_data = [
    ["Number Input", "Date Input", "List Selection", "Text Length"],
    [25, "2023-01-15", "Option A", "Sample"],
    [30, "2023-02-20", "Option B", "Text"],
    ["", "", "", ""]
]

for row_idx, row_data in enumerate(validation_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        ws2.cell(row=row_idx, column=col_idx, value=value)

# Add data validation rules
# 1. Number validation (whole numbers between 1-100)
number_validation = DataValidation(
    type="whole",
    operator="between", 
    formula1=1,
    formula2=100,
    showErrorMessage=True,
    errorTitle="Invalid Number",
    error="Please enter a number between 1 and 100"
)
ws2.add_data_validation(number_validation)
number_validation.add("A2:A10")

# 2. Date validation
date_validation = DataValidation(
    type="date",
    operator="greaterThan",
    formula1="2020-01-01",
    showErrorMessage=True,
    errorTitle="Invalid Date", 
    error="Date must be after 2020-01-01"
)
ws2.add_data_validation(date_validation)
date_validation.add("B2:B10")

# 3. List validation
list_validation = DataValidation(
    type="list",
    formula1='"Option A,Option B,Option C"',
    showErrorMessage=True,
    errorTitle="Invalid Selection",
    error="Please select from the dropdown list"
)
ws2.add_data_validation(list_validation)
list_validation.add("C2:C10")

# 4. Text length validation
text_validation = DataValidation(
    type="textLength",
    operator="between",
    formula1=1,
    formula2=50,
    showErrorMessage=True,
    errorTitle="Text Too Long",
    error="Text must be between 1 and 50 characters"
)
ws2.add_data_validation(text_validation)
text_validation.add("D2:D10")

# Sheet 3: More named ranges (sheet-specific)
ws3 = wb.create_sheet("Summary")
ws3["A1"] = "Summary Data"
ws3["A2"] = "Total Sales"
ws3["B2"] = "=SUM(SalesColumn)"

# Add sheet-specific named range
wb.defined_names["SummaryTitle"] = DefinedName("SummaryTitle", attr_text="Summary.A1", localSheetId=2)

# Save the workbook
filename = "comprehensive_test.xlsx"
wb.save(filename)
wb.close()

print(f"✓ Created {filename} with named ranges and validation rules")

print("\n" + "=" * 80)
print("1. XLSX_SHEET_NAMES - List all worksheet names")
print("=" * 80)

sheet_names = xlsx_sheet_names(filename)
print(f"Sheet names ({len(sheet_names)} total):")
for i, name in enumerate(sheet_names, 1):
    print(f"  {i}. {name}")

print("\n" + "=" * 80)
print("2. XLSX_NAMES - Extract named ranges and formulas")
print("=" * 80)

names_df = xlsx_names(filename)
print(f"Named ranges found: {len(names_df)}")

if len(names_df) > 0:
    print("\nNamed ranges details:")
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    print(names_df.to_string(index=False))
    
    print(f"\nBreakdown:")
    print(f"  Global ranges: {len(names_df[names_df['sheet'].isna()])}")
    print(f"  Sheet-specific ranges: {len(names_df[names_df['sheet'].notna()])}")
    print(f"  Cell ranges: {len(names_df[names_df['is_range'] == True])}")
    print(f"  Complex formulas: {len(names_df[names_df['is_range'] == False])}")
else:
    print("No named ranges found in this file")

print("\n" + "=" * 80)
print("3. XLSX_VALIDATION - Extract data validation rules")
print("=" * 80)

validation_df = xlsx_validation(filename)
print(f"Validation rules found: {len(validation_df)}")

if len(validation_df) > 0:
    print("\nValidation rules details:")
    
    # Show key columns for readability
    key_columns = ['sheet', 'ref', 'type', 'operator', 'formula1', 'formula2', 'error_title']
    available_columns = [col for col in key_columns if col in validation_df.columns]
    
    print(validation_df[available_columns].to_string(index=False))
    
    print(f"\nValidation types found:")
    type_counts = validation_df['type'].value_counts()
    for vtype, count in type_counts.items():
        print(f"  {vtype}: {count} rules")
        
    print(f"\nValidation by sheet:")
    sheet_counts = validation_df['sheet'].value_counts()
    for sheet, count in sheet_counts.items():
        print(f"  {sheet}: {count} rules")
else:
    print("No validation rules found in this file")

print("\n" + "=" * 80)
print("4. COMPARISON WITH XLSX_CELLS")
print("=" * 80)

# Show how these functions complement xlsx_cells
print("Reading specific sheet using xlsx_sheet_names result:")
first_sheet = sheet_names[0]
cells_df = xlsx_cells(filename, sheets=first_sheet)

print(f"Cells in '{first_sheet}': {len(cells_df)} total")
content_cells = cells_df[~cells_df['is_blank']]
print(f"Cells with content: {len(content_cells)}")

print(f"\nFirst few cells from {first_sheet}:")
print(content_cells[['address', 'content', 'data_type']].head().to_string(index=False))

print("\n" + "=" * 80)
print("5. PRACTICAL USAGE PATTERNS")
print("=" * 80)

print("# Get all sheet names first")
print("sheets = xlsx_sheet_names('file.xlsx')")
print()

print("# Check for named ranges before processing")
print("names = xlsx_names('file.xlsx')")
print("if len(names) > 0:")
print("    print('Found named ranges:', names['name'].tolist())")
print()

print("# Check validation rules to understand data constraints")
print("validation = xlsx_validation('file.xlsx')")
print("if len(validation) > 0:")
print("    print('Validation types:', validation['type'].unique())")
print()

print("# Then read cell data with full context")
print("cells = xlsx_cells('file.xlsx')")

print("\n" + "=" * 80)
print("6. FUNCTION SIGNATURES SUMMARY")
print("=" * 80)

print("""
xlsx_sheet_names(path, check_filetype=True) -> List[str]
    Returns list of worksheet names in order

xlsx_names(path, check_filetype=True) -> pd.DataFrame  
    Returns named ranges with columns: sheet, name, formula, comment, hidden, is_range

xlsx_validation(path, sheets=None, check_filetype=True) -> pd.DataFrame
    Returns validation rules with columns: sheet, ref, type, operator, formula1, 
    formula2, allow_blank, show_input_message, show_error_message, prompt_title, 
    prompt, error_title, error, error_style

xlsx_cells(path, sheets=None, check_filetype=True, include_blank_cells=True) -> pd.DataFrame
    Returns all cell data in tidy format (23 columns as shown previously)

xlsx_formats(path) -> Dict[str, Any]
    Returns formatting information (fonts, fills, borders, number_formats)
""")

print("All functions now match the R tidyxl package behavior exactly!")

print("\n" + "=" * 80)
print("DEMONSTRATION COMPLETE!")
print("=" * 80)