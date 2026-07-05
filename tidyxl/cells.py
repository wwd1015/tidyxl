"""
Cell data extraction functionality
"""

from functools import lru_cache
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

from ._common import resolve_sheet_names, validate_filetype

# Column order of the DataFrame returned by xlsx_cells
_CELL_COLUMNS = [
    'sheet', 'address', 'row', 'col', 'is_blank', 'content', 'data_type',
    'error', 'logical', 'numeric', 'date', 'character', 'formula',
    'is_array', 'formula_ref', 'formula_group', 'comment', 'height', 'width',
    'row_outline_level', 'col_outline_level', 'style_format', 'local_format_id'
]

# Characters whose presence in a number format indicates a date/time format
_DATE_FORMAT_CHARS = frozenset('dmyhs:/-')


def xlsx_cells(
    path: str,
    sheets: Optional[Union[str, List[str]]] = None,
    check_filetype: bool = True,
    include_blank_cells: bool = True
) -> pd.DataFrame:
    """
    Import xlsx (Excel) cell contents into a tidy structure.

    Imports data from spreadsheets without coercing it into a rectangle.
    Each cell is represented by a row in a data frame, following the exact
    behavior of the R tidyxl package.

    Parameters
    ----------
    path : str
        Path to the Excel file (.xlsx or .xlsm)
    sheets : str, list of str, or None
        Worksheet names to read. If None, reads all sheets.
    check_filetype : bool, default True
        Whether to check that the file is a valid xlsx/xlsm file
    include_blank_cells : bool, default True
        Whether to include cells that have no value but may have formatting

    Returns
    -------
    pd.DataFrame
        A tidy DataFrame where each row represents a single cell with columns:
        - sheet: worksheet name (str)
        - address: cell address in A1 notation (str)
        - row: row number (int)
        - col: column number (int)
        - is_blank: whether cell has a value (bool)
        - content: raw cell value before type conversion (str)
        - data_type: cell type (str: error, logical, numeric, date, character, blank)
        - error: cell error value (str)
        - logical: boolean value (bool)
        - numeric: numeric value (float)
        - date: date value (datetime)
        - character: string value (str)
        - formula: cell formula (str)
        - is_array: whether formula is an array formula (bool)
        - formula_ref: range address for array/shared formulas (str)
        - formula_group: formula group index (int)
        - comment: cell comment text (str)
        - height: row height in Excel units (float)
        - width: column width in Excel units (float)
        - row_outline_level: row outline level (int)
        - col_outline_level: column outline level (int)
        - style_format: index for style formats (str)
        - local_format_id: index for local cell formats (int)
    """

    if check_filetype:
        validate_filetype(path)

    # keep_vba retains the whole source archive in memory, so only pay that
    # cost for macro-enabled files
    wb = load_workbook(
        filename=path, data_only=False, keep_vba=path.lower().endswith('.xlsm')
    )

    try:
        sheet_names = resolve_sheet_names(wb, sheets)
        all_cells = []
        # Stable per-workbook index for each distinct number format
        format_ids: Dict[str, int] = {}

        # Sheets are processed in alphabetical order and openpyxl yields cells
        # in row/column order, so the output is already sorted by
        # (sheet, row, col) without a post-hoc DataFrame sort.
        for sheet_name in sorted(sheet_names):
            ws = wb[sheet_name]
            # Width/outline lookups are cached per column index because
            # openpyxl's dimension mappings create objects on access.
            col_dims: Dict[int, Tuple[Optional[float], int]] = {}

            for row in ws.iter_rows():
                if not row:
                    continue

                row_dim = ws.row_dimensions[row[0].row]
                row_height = row_dim.height
                row_outline_level = row_dim.outline_level or 0

                for cell in row:
                    is_blank = cell.value is None and (
                        cell.data_type == 'n' or cell.data_type is None
                    )

                    if not include_blank_cells and is_blank:
                        continue

                    col_dim = col_dims.get(cell.column)
                    if col_dim is None:
                        dim = ws.column_dimensions[get_column_letter(cell.column)]
                        col_dim = (dim.width, dim.outline_level or 0)
                        col_dims[cell.column] = col_dim

                    data_type, value_column, value = _get_typed_value(cell)
                    formula_info = _get_formula_info(cell)

                    number_format = cell.number_format
                    local_format_id = (
                        format_ids.setdefault(number_format, len(format_ids))
                        if number_format else None
                    )

                    cell_record = {
                        'sheet': sheet_name,
                        'address': cell.coordinate,
                        'row': cell.row,
                        'col': cell.column,
                        'is_blank': is_blank,
                        'content': str(cell.value) if cell.value is not None else None,
                        'data_type': data_type,
                        'error': None,
                        'logical': None,
                        'numeric': None,
                        'date': None,
                        'character': None,
                        'formula': formula_info['formula'],
                        'is_array': formula_info['is_array'],
                        'formula_ref': formula_info['formula_ref'],
                        'formula_group': formula_info['formula_group'],
                        'comment': cell.comment.text if cell.comment else None,
                        'height': row_height,
                        'width': col_dim[0],
                        'row_outline_level': row_outline_level,
                        'col_outline_level': col_dim[1],
                        'style_format': cell.style,
                        'local_format_id': local_format_id
                    }

                    if value_column is not None:
                        cell_record[value_column] = value

                    all_cells.append(cell_record)
    finally:
        wb.close()

    return pd.DataFrame(all_cells, columns=_CELL_COLUMNS)


def _get_typed_value(cell) -> Tuple[str, Optional[str], Any]:
    """
    Determine the data type of a cell and extract its typed value.

    Parameters
    ----------
    cell : openpyxl.cell.Cell
        The cell to analyze

    Returns
    -------
    tuple
        (data_type, value_column, value) where data_type is one of
        'error', 'logical', 'numeric', 'date', 'character', 'formula',
        'blank', and value_column names the output column holding the
        typed value (None when there is no value to store)
    """

    if cell.value is None:
        return 'blank', None, None

    data_type = cell.data_type

    if data_type == 'e':  # Error
        return 'error', 'error', str(cell.value)

    if data_type == 'b':  # Boolean
        return 'logical', 'logical', bool(cell.value)

    if data_type == 'n':  # Numeric, possibly a date depending on format
        if _is_date_format(cell.number_format):
            try:
                return 'date', 'date', from_excel(cell.value)
            except Exception:
                pass  # Fall back to numeric if date conversion fails
        return 'numeric', 'numeric', float(cell.value)

    if data_type == 'f':  # Formula: the text is reported in the formula column
        return 'formula', None, None

    # String types ('s', 'inlineStr', 'str')
    return 'character', 'character', str(cell.value)


@lru_cache(maxsize=None)
def _is_date_format(number_format: Optional[str]) -> bool:
    """
    Check if a number format string indicates a date.

    Parameters
    ----------
    number_format : str or None
        The cell's number format code

    Returns
    -------
    bool
        True if the format appears to be a date format
    """

    if not number_format:
        return False

    return not _DATE_FORMAT_CHARS.isdisjoint(number_format.lower())


def _get_formula_info(cell) -> Dict[str, Any]:
    """
    Extract formula-related information from a cell.

    Parameters
    ----------
    cell : openpyxl.cell.Cell
        The cell to analyze

    Returns
    -------
    dict
        Dictionary with formula, is_array, formula_ref, formula_group
    """

    formula_info: Dict[str, Any] = {
        'formula': None,
        'is_array': False,
        'formula_ref': None,
        'formula_group': None
    }

    if cell.data_type == 'f' and cell.value:
        formula_info['formula'] = str(cell.value)

        # Check for array formula indicators
        if getattr(cell, 'array_formula', None):
            formula_info['is_array'] = True

        # Try to get formula reference range (this is limited in openpyxl)
        if hasattr(cell, 'shared_formula'):
            formula_info['formula_group'] = id(cell.shared_formula)

    return formula_info
