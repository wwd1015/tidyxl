"""
Workbook-level functionality for sheet names and metadata
"""

import re
from typing import List

import pandas as pd
from openpyxl import load_workbook

from ._common import validate_filetype

# Leading sheet reference (e.g. "Sheet1!") in a defined-name formula
_SHEET_PREFIX_RE = re.compile(r'^[^!]*!')

# Single cell or rectangular range, with optional absolute ($) markers
_CELL_RANGE_RE = re.compile(r'^\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?$')


def xlsx_sheet_names(path: str, check_filetype: bool = True) -> List[str]:
    """
    List sheets in an xlsx (Excel) file.

    Returns the names of the sheets in a workbook, as a list of strings,
    in the order they appear when opening the spreadsheet.

    Parameters
    ----------
    path : str
        Path to the Excel file (.xlsx or .xlsm)
    check_filetype : bool, default True
        Whether to check that the file is a valid xlsx/xlsm file

    Returns
    -------
    List[str]
        List of worksheet names in their original order
    """

    if check_filetype:
        validate_filetype(path)

    # Load workbook (read-only for efficiency)
    wb = load_workbook(filename=path, read_only=True, data_only=False)

    try:
        return wb.sheetnames
    finally:
        wb.close()


def xlsx_names(path: str, check_filetype: bool = True) -> pd.DataFrame:
    """
    Import named formulas from xlsx (Excel) files.

    Extracts named ranges and named formulas (defined names) from Excel files,
    including both global and sheet-specific named ranges.

    Parameters
    ----------
    path : str
        Path to the Excel file (.xlsx or .xlsm)
    check_filetype : bool, default True
        Whether to check that the file is a valid xlsx/xlsm file

    Returns
    -------
    pd.DataFrame
        A DataFrame with columns:
        - sheet: Sheet name (None if globally defined)
        - name: Name of the formula/range
        - formula: Cell range or formula definition
        - comment: Description by spreadsheet author
        - hidden: Visibility status
        - is_range: Whether formula represents a cell range
    """

    if check_filetype:
        validate_filetype(path)

    wb = load_workbook(filename=path, data_only=False)

    names_list = []

    try:
        sheet_names = wb.sheetnames

        for name, defined_name in wb.defined_names.items():
            # Determine if it's sheet-specific or global
            sheet_name = None
            local_sheet_id = getattr(defined_name, 'localSheetId', None)
            if local_sheet_id is not None and local_sheet_id < len(sheet_names):
                sheet_name = sheet_names[local_sheet_id]

            formula_text = str(getattr(defined_name, 'attr_text', None) or "")

            names_list.append({
                'sheet': sheet_name,
                'name': name,
                'formula': formula_text,
                'comment': getattr(defined_name, 'comment', None),
                'hidden': getattr(defined_name, 'hidden', False),
                'is_range': _is_cell_range(formula_text)
            })

    finally:
        wb.close()

    df = pd.DataFrame(
        names_list,
        columns=['sheet', 'name', 'formula', 'comment', 'hidden', 'is_range']
    )

    # Sort by sheet (global names first), then by name
    if not df.empty:
        df['_sort_key'] = df['sheet'].fillna('')
        df = df.sort_values(['_sort_key', 'name']).drop('_sort_key', axis=1).reset_index(drop=True)

    return df


def _is_cell_range(formula_text: str) -> bool:
    """
    Check if a formula represents a cell range vs a complex formula.

    Parameters
    ----------
    formula_text : str
        The formula text to analyze

    Returns
    -------
    bool
        True if it appears to be a simple cell range reference
    """

    if not formula_text:
        return False

    # Remove sheet references for analysis
    clean_formula = _SHEET_PREFIX_RE.sub('', formula_text)

    return bool(_CELL_RANGE_RE.match(clean_formula.strip()))
