"""
Core functionality for tidyxl package
"""

import re
from typing import Any, Dict, List, Optional, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def xlsx_cells(
    path: str,
    sheets: Optional[Union[str, List[str]]] = None,
    check_filetype: bool = True,
    include_blank_cells: bool = True
) -> pd.DataFrame:
    """
    Import xlsx (Excel) cell contents into a tidy structure.

    Imports data from spreadsheets without coercing it into a rectangle.
    Each cell is represented by a row in a data frame, following the exact
    behavior of the R tidyxl package.

    Parameters
    ----------
    path : str
        Path to the Excel file (.xlsx or .xlsm)
    sheets : str, list of str, or None
        Worksheet names to read. If None, reads all sheets.
    check_filetype : bool, default True
        Whether to check that the file is a valid xlsx/xlsm file
    include_blank_cells : bool, default True
        Whether to include cells that have no value but may have formatting

    Returns
    -------
    pd.DataFrame
        A tidy DataFrame where each row represents a single cell with columns:
        - sheet: worksheet name (str)
        - address: cell address in A1 notation (str)
        - row: row number (int)
        - col: column number (int)
        - is_blank: whether cell has a value (bool)
        - content: raw cell value before type conversion (str)
        - data_type: cell type (str: error, logical, numeric, date, character, blank)
        - error: cell error value (str)
        - logical: boolean value (bool)
        - numeric: numeric value (float)
        - date: date value (datetime)
        - character: string value (str)
        - formula: cell formula (str)
        - is_array: whether formula is an array formula (bool)
        - formula_ref: range address for array/shared formulas (str)
        - formula_group: formula group index (int)
        - comment: cell comment text (str)
        - height: row height in Excel units (float)
        - width: column width in Excel units (float)
        - row_outline_level: row outline level (int)
        - col_outline_level: column outline level (int)
        - style_format: index for style formats (str)
        - local_format_id: index for local cell formats (int)
    """

    # Check file type if requested
    if check_filetype:
        if not path.lower().endswith(('.xlsx', '.xlsm')):
            raise ValueError("File must be .xlsx or .xlsm format")

    # Load workbook
    wb = load_workbook(filename=path, data_only=False, keep_vba=True)

    # Determine which sheets to process
    if sheets is None:
        sheet_names = wb.sheetnames
    elif isinstance(sheets, str):
        sheet_names = [sheets]
    else:
        sheet_names = sheets

    # Validate sheet names
    available_sheets = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name not in available_sheets:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}")

    all_cells = []

    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        # Get all cells in the worksheet
        for row in ws.iter_rows():
            for cell in row:
                # Determine if cell is blank
                is_blank = cell.value is None and (cell.data_type == 'n' or cell.data_type is None)

                # Skip blank cells if not requested
                if not include_blank_cells and is_blank:
                    continue

                # Get raw content as string
                content = str(cell.value) if cell.value is not None else None

                # Determine data type and extract typed values
                data_type, typed_values = _get_cell_data_and_values(cell)

                # Get formula information
                formula_info = _get_formula_info(cell)

                # Get comment
                comment = cell.comment.text if cell.comment else None

                # Get dimensions
                row_height = ws.row_dimensions[cell.row].height
                col_width = ws.column_dimensions[get_column_letter(cell.column)].width

                # Get outline levels
                row_outline_level = ws.row_dimensions[cell.row].outline_level or 0
                col_outline_level = ws.column_dimensions[get_column_letter(cell.column)].outline_level or 0

                # Create cell record matching R tidyxl structure
                cell_record = {
                    'sheet': sheet_name,
                    'address': cell.coordinate,
                    'row': cell.row,
                    'col': cell.column,
                    'is_blank': is_blank,
                    'content': content,
                    'data_type': data_type,
                    'error': typed_values.get('error'),
                    'logical': typed_values.get('logical'),
                    'numeric': typed_values.get('numeric'),
                    'date': typed_values.get('date'),
                    'character': typed_values.get('character'),
                    'formula': formula_info['formula'],
                    'is_array': formula_info['is_array'],
                    'formula_ref': formula_info['formula_ref'],
                    'formula_group': formula_info['formula_group'],
                    'comment': comment,
                    'height': row_height,
                    'width': col_width,
                    'row_outline_level': row_outline_level,
                    'col_outline_level': col_outline_level,
                    'style_format': cell.style if hasattr(cell, 'style') else None,
                    'local_format_id': id(cell.number_format) if cell.number_format else None
                }

                all_cells.append(cell_record)

    # Convert to DataFrame
    df = pd.DataFrame(all_cells)

    # Sort by sheet, row, column for consistent output
    if not df.empty:
        df = df.sort_values(['sheet', 'row', 'col']).reset_index(drop=True)

    return df


def xlsx_formats(path: str) -> Dict[str, Any]:
    """
    Import xlsx (Excel) formatting information.

    This function extracts formatting information from Excel files,
    providing details about fonts, fills, borders, and number formats.

    Parameters
    ----------
    path : str
        Path to the Excel file (.xlsx or .xlsm)

    Returns
    -------
    dict
        Dictionary containing formatting information with keys:
        - fonts: font formatting details
        - fills: fill/background formatting details
        - borders: border formatting details
        - number_formats: number format details
    """

    wb = load_workbook(filename=path, data_only=False)

    formats = {
        'fonts': [],
        'fills': [],
        'borders': [],
        'number_formats': []
    }

    # Extract font information
    if hasattr(wb, '_fonts'):
        for font in wb._fonts:
            font_info = {
                'name': font.name,
                'size': font.size,
                'bold': font.bold,
                'italic': font.italic,
                'underline': font.underline,
                'color': str(font.color.rgb) if font.color and hasattr(font.color, 'rgb') else None
            }
            formats['fonts'].append(font_info)

    # Extract fill information
    if hasattr(wb, '_fills'):
        for fill in wb._fills:
            fill_info = {
                'fill_type': fill.fill_type,
                'start_color': str(fill.start_color.rgb) if hasattr(fill.start_color, 'rgb') else None,
                'end_color': str(fill.end_color.rgb) if hasattr(fill.end_color, 'rgb') else None
            }
            formats['fills'].append(fill_info)

    # Extract border information
    if hasattr(wb, '_borders'):
        for border in wb._borders:
            border_info = {
                'left': str(border.left.style) if border.left else None,
                'right': str(border.right.style) if border.right else None,
                'top': str(border.top.style) if border.top else None,
                'bottom': str(border.bottom.style) if border.bottom else None
            }
            formats['borders'].append(border_info)

    # Extract number format information
    if hasattr(wb, '_number_formats'):
        for num_format in wb._number_formats:
            formats['number_formats'].append({
                'format_code': num_format.format_code,
                'format_id': num_format.format_id
            })

    return formats


def _get_cell_data_and_values(cell) -> tuple[str, Dict[str, Any]]:
    """
    Determine the data type of a cell and extract typed values.

    Parameters
    ----------
    cell : openpyxl.cell.Cell
        The cell to analyze

    Returns
    -------
    tuple
        (data_type, typed_values_dict) where data_type is one of:
        'error', 'logical', 'numeric', 'date', 'character', 'blank'
        and typed_values_dict contains the appropriate typed value
    """

    typed_values = {
        'error': None,
        'logical': None,
        'numeric': None,
        'date': None,
        'character': None
    }

    if cell.value is None:
        return 'blank', typed_values

    # Handle different openpyxl data types
    if cell.data_type == 'e':  # Error
        typed_values['error'] = str(cell.value)
        return 'error', typed_values

    elif cell.data_type == 'b':  # Boolean
        typed_values['logical'] = bool(cell.value)
        return 'logical', typed_values

    elif cell.data_type == 'n':  # Numeric
        # Check if it's a date by looking at number format
        if _is_date_format(cell):
            try:
                # Convert Excel date serial to datetime
                from openpyxl.utils.datetime import from_excel
                typed_values['date'] = from_excel(cell.value)
                return 'date', typed_values
            except Exception:
                # Fall back to numeric if date conversion fails
                typed_values['numeric'] = float(cell.value)
                return 'numeric', typed_values
        else:
            typed_values['numeric'] = float(cell.value)
            return 'numeric', typed_values

    elif cell.data_type == 'f':  # Formula - don't return formula as data_type
        # For formulas, we need to determine the result type
        # Since we can't evaluate formulas, we'll treat as character
        typed_values['character'] = str(cell.value) if cell.value else None
        return 'character', typed_values

    else:  # String types ('s', 'inlineStr', 'str')
        typed_values['character'] = str(cell.value)
        return 'character', typed_values


def _is_date_format(cell) -> bool:
    """
    Check if a cell's number format indicates it's a date.

    Parameters
    ----------
    cell : openpyxl.cell.Cell
        The cell to check

    Returns
    -------
    bool
        True if the cell appears to be formatted as a date
    """

    if not cell.number_format:
        return False

    # Common date format indicators
    date_indicators = ['d', 'm', 'y', 'h', 's', ':', '/', '-']
    format_str = cell.number_format.lower()

    # Check if format contains date indicators
    return any(indicator in format_str for indicator in date_indicators)


def _get_formula_info(cell) -> Dict[str, Any]:
    """
    Extract formula-related information from a cell.

    Parameters
    ----------
    cell : openpyxl.cell.Cell
        The cell to analyze

    Returns
    -------
    dict
        Dictionary with formula, is_array, formula_ref, formula_group
    """

    formula_info = {
        'formula': None,
        'is_array': False,
        'formula_ref': None,
        'formula_group': None
    }

    if cell.data_type == 'f' and cell.value:
        formula_info['formula'] = str(cell.value)

        # Check for array formula indicators
        if hasattr(cell, 'array_formula') and cell.array_formula:
            formula_info['is_array'] = True

        # Try to get formula reference range (this is limited in openpyxl)
        if hasattr(cell, 'shared_formula'):
            formula_info['formula_group'] = id(cell.shared_formula)

    return formula_info


def xlsx_sheet_names(path: str, check_filetype: bool = True) -> List[str]:
    """
    List sheets in an xlsx (Excel) file.

    Returns the names of the sheets in a workbook, as a list of strings,
    in the order they appear when opening the spreadsheet.

    Parameters
    ----------
    path : str
        Path to the Excel file (.xlsx or .xlsm)
    check_filetype : bool, default True
        Whether to check that the file is a valid xlsx/xlsm file

    Returns
    -------
    List[str]
        List of worksheet names in their original order
    """

    # Check file type if requested
    if check_filetype:
        if not path.lower().endswith(('.xlsx', '.xlsm')):
            raise ValueError("File must be .xlsx or .xlsm format")

    # Load workbook (read-only for efficiency)
    wb = load_workbook(filename=path, read_only=True, data_only=False)

    try:
        return wb.sheetnames
    finally:
        wb.close()


def xlsx_names(path: str, check_filetype: bool = True) -> pd.DataFrame:
    """
    Import named formulas from xlsx (Excel) files.

    Extracts named ranges and named formulas (defined names) from Excel files,
    including both global and sheet-specific named ranges.

    Parameters
    ----------
    path : str
        Path to the Excel file (.xlsx or .xlsm)
    check_filetype : bool, default True
        Whether to check that the file is a valid xlsx/xlsm file

    Returns
    -------
    pd.DataFrame
        A DataFrame with columns:
        - sheet: Sheet name (None if globally defined)
        - name: Name of the formula/range
        - formula: Cell range or formula definition
        - comment: Description by spreadsheet author
        - hidden: Visibility status
        - is_range: Whether formula represents a cell range
    """

    # Check file type if requested
    if check_filetype:
        if not path.lower().endswith(('.xlsx', '.xlsm')):
            raise ValueError("File must be .xlsx or .xlsm format")

    # Load workbook
    wb = load_workbook(filename=path, data_only=False)

    names_list = []

    try:
        # Get defined names from workbook
        for name, defined_name in wb.defined_names.items():
            # Determine if it's sheet-specific or global
            sheet_name = None
            if hasattr(defined_name, 'localSheetId') and defined_name.localSheetId is not None:
                # Get sheet name by index
                sheet_names = wb.sheetnames
                if defined_name.localSheetId < len(sheet_names):
                    sheet_name = sheet_names[defined_name.localSheetId]

            # Check if it's a range (contains cell references) vs formula
            formula_text = str(defined_name.attr_text) if hasattr(defined_name, 'attr_text') and defined_name.attr_text else ""
            is_range = _is_cell_range(formula_text)

            name_record = {
                'sheet': sheet_name,
                'name': name,
                'formula': formula_text,
                'comment': getattr(defined_name, 'comment', None),
                'hidden': getattr(defined_name, 'hidden', False),
                'is_range': is_range
            }

            names_list.append(name_record)

    finally:
        wb.close()

    # Convert to DataFrame
    df = pd.DataFrame(names_list)

    # Sort by sheet (global first), then by name
    if not df.empty:
        df['_sort_key'] = df['sheet'].fillna('')  # Global names first
        df = df.sort_values(['_sort_key', 'name']).drop('_sort_key', axis=1).reset_index(drop=True)

    return df


def xlsx_validation(
    path: str,
    sheets: Optional[Union[str, List[str]]] = None,
    check_filetype: bool = True
) -> pd.DataFrame:
    """
    Import data validation rules of cells in xlsx (Excel) files.

    Extracts data validation rules from Excel cells, including numeric ranges,
    date constraints, list restrictions, and custom formula-driven rules.

    Parameters
    ----------
    path : str
        Path to the Excel file (.xlsx or .xlsm)
    sheets : str, list of str, or None
        Worksheet names to read. If None, reads all sheets.
    check_filetype : bool, default True
        Whether to check that the file is a valid xlsx/xlsm file

    Returns
    -------
    pd.DataFrame
        A DataFrame with columns:
        - sheet: Worksheet with validation rule
        - ref: Cell addresses with rules (e.g., 'A1:A10')
        - type: Data validation type (whole, decimal, list, date, time, textLength, custom)
        - operator: Comparison operator (between, equal, notEqual, greaterThan, etc.)
        - formula1: First validation criterion
        - formula2: Second validation criterion (for between/notBetween)
        - allow_blank: Whether blank cells are allowed
        - show_input_message: Whether to show input message
        - show_error_message: Whether to show error message
        - prompt_title: Input message title
        - prompt: Input message text
        - error_title: Error message title
        - error: Error message text
        - error_style: Error style (stop, warning, information)
    """

    # Check file type if requested
    if check_filetype:
        if not path.lower().endswith(('.xlsx', '.xlsm')):
            raise ValueError("File must be .xlsx or .xlsm format")

    # Load workbook
    wb = load_workbook(filename=path, data_only=False)

    # Determine which sheets to process
    if sheets is None:
        sheet_names = wb.sheetnames
    elif isinstance(sheets, str):
        sheet_names = [sheets]
    else:
        sheet_names = sheets

    # Validate sheet names
    available_sheets = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name not in available_sheets:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}")

    validation_list = []

    try:
        for sheet_name in sheet_names:
            ws = wb[sheet_name]

            # Get data validation rules
            if hasattr(ws, 'data_validations'):
                for dv in ws.data_validations.dataValidation:
                    validation_record = {
                        'sheet': sheet_name,
                        'ref': str(dv.sqref) if dv.sqref else None,
                        'type': dv.type,
                        'operator': dv.operator,
                        'formula1': dv.formula1,
                        'formula2': dv.formula2,
                        'allow_blank': dv.allowBlank,
                        'show_input_message': dv.showInputMessage,
                        'show_error_message': dv.showErrorMessage,
                        'prompt_title': dv.promptTitle,
                        'prompt': dv.prompt,
                        'error_title': dv.errorTitle,
                        'error': dv.error,
                        'error_style': dv.errorStyle
                    }

                    validation_list.append(validation_record)

    finally:
        wb.close()

    # Convert to DataFrame
    df = pd.DataFrame(validation_list)

    # Sort by sheet, then by ref
    if not df.empty:
        df = df.sort_values(['sheet', 'ref']).reset_index(drop=True)

    return df


def _is_cell_range(formula_text: str) -> bool:
    """
    Check if a formula represents a cell range vs a complex formula.

    Parameters
    ----------
    formula_text : str
        The formula text to analyze

    Returns
    -------
    bool
        True if it appears to be a simple cell range reference
    """

    if not formula_text:
        return False

    # Remove sheet references for analysis
    clean_formula = re.sub(r'^[^!]*!', '', formula_text)

    # Simple patterns that indicate cell ranges
    range_patterns = [
        r'^[A-Z]+\d+$',  # Single cell (e.g., A1)
        r'^[A-Z]+\d+:[A-Z]+\d+$',  # Range (e.g., A1:B10)
        r'^\$?[A-Z]+\$?\d+$',  # Absolute single cell
        r'^\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+$',  # Absolute range
    ]

    return any(re.match(pattern, clean_formula.strip()) for pattern in range_patterns)
