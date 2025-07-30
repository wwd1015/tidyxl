"""
Pytest configuration and shared fixtures for tidyxl tests
"""

import tempfile
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation


@pytest.fixture
def temp_dir():
    """Create a temporary directory for test files"""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def sample_excel_file(temp_dir):
    """Create a sample Excel file with various data types"""
    filepath = temp_dir / "sample.xlsx"

    # Create workbook with sample data
    wb = Workbook()
    ws = wb.active
    ws.title = "TestData"

    # Add headers
    headers = ["Text", "Number", "Boolean", "Date", "Formula"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Add data rows
    data = [
        ["Hello", 42, True, datetime(2023, 1, 15), "=B2*2"],
        ["World", 3.14, False, datetime(2023, 6, 1), "=B3+10"],
        ["Test", -100, True, datetime(2023, 12, 25), "=SUM(B2:B4)"]
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Add a comment to one cell
            if row_idx == 2 and col_idx == 1:
                from openpyxl.comments import Comment
                cell.comment = Comment("This is a test comment", "Test Author")

    # Save the workbook
    wb.save(filepath)
    wb.close()

    return str(filepath)


@pytest.fixture
def multi_sheet_excel_file(temp_dir):
    """Create an Excel file with multiple sheets"""
    filepath = temp_dir / "multi_sheet.xlsx"

    wb = Workbook()

    # Sheet 1: Employee data
    ws1 = wb.active
    ws1.title = "Employees"
    emp_data = [
        ["Name", "Age", "Department"],
        ["Alice", 25, "Engineering"],
        ["Bob", 30, "Sales"],
        ["Charlie", 35, "Marketing"]
    ]
    for row_idx, row_data in enumerate(emp_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)

    # Sheet 2: Product data
    ws2 = wb.create_sheet("Products")
    prod_data = [
        ["Product", "Price", "Stock"],
        ["Widget A", 10.99, 100],
        ["Widget B", 15.50, 50]
    ]
    for row_idx, row_data in enumerate(prod_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)

    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "Total Employees"
    ws3["B1"] = "=COUNTA(Employees.A:A)-1"

    wb.save(filepath)
    wb.close()

    return str(filepath)


@pytest.fixture
def excel_with_named_ranges(temp_dir):
    """Create an Excel file with named ranges"""
    filepath = temp_dir / "named_ranges.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Add some data
    data = [
        ["Item", "Value"],
        ["A", 10],
        ["B", 20],
        ["C", 30]
    ]
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Add named ranges
    wb.defined_names["DataRange"] = DefinedName("DataRange", attr_text="Data.A1:B4")
    wb.defined_names["ValueColumn"] = DefinedName("ValueColumn", attr_text="Data.B2:B4")

    wb.save(filepath)
    wb.close()

    return str(filepath)


@pytest.fixture
def excel_with_validation(temp_dir):
    """Create an Excel file with data validation rules"""
    filepath = temp_dir / "validation.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "ValidationSheet"

    # Add headers
    headers = ["Number", "Date", "List", "Text"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Add validation rules
    # Number validation
    number_val = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=100,
        showErrorMessage=True,
        errorTitle="Invalid Number"
    )
    ws.add_data_validation(number_val)
    number_val.add("A2:A10")

    # List validation
    list_val = DataValidation(
        type="list",
        formula1='"Option1,Option2,Option3"'
    )
    ws.add_data_validation(list_val)
    list_val.add("C2:C10")

    wb.save(filepath)
    wb.close()

    return str(filepath)


@pytest.fixture
def empty_excel_file(temp_dir):
    """Create an empty Excel file"""
    filepath = temp_dir / "empty.xlsx"

    wb = Workbook()
    wb.save(filepath)
    wb.close()

    return str(filepath)
